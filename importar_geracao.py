import os
import psycopg2
import openpyxl
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path
import re
from openpyxl.utils.datetime import from_excel # Importa a função de conversão de datas do openpyxl

CAMINHO_ARQUIVO = Path(__file__).parent / "planilha.xlsx"

def conectar():
    conn = psycopg2.connect(
        host="ep-cold-waterfall-acdcoi56-pooler.sa-east-1.aws.neon.tech",
        dbname="neondb",
        user="neondb_owner",
        password="npg_8loV6RANevUS",
        port="5432",
    )
    conn.autocommit = True
    return conn

def normalizar_nome_para_comparacao(nome):
    """
    Normaliza o nome removendo espaços em branco, caracteres de largura zero,
    acentos e convertendo para minúsculas para comparação.
    """
    if not isinstance(nome, str):
        return ""
    nome = nome.replace("\u200b", "")
    # Decompõe caracteres e remove marcas diacríticas (acentos)
    nome = unicodedata.normalize("NFKD", nome).encode("ascii", "ignore").decode("utf-8")
    return nome.strip().lower()

def carregar_usinas_do_db(cursor):
    """
    Carrega todos os nomes de usinas do banco de dados e seus IDs,
    normalizando os nomes para comparação.
    Retorna um dicionário {nome_normalizado: id_usina}.
    """
    usinas_map = {}
    try:
        cursor.execute('SELECT id, nome FROM "Usina"')
        for usina_id, nome_db in cursor.fetchall():
            nome_normalizado = normalizar_nome_para_comparacao(nome_db)
            usinas_map[nome_normalizado] = usina_id
        print(f"[DEBUG] Usinas carregadas do DB ({len(usinas_map)}): {list(usinas_map.keys())[:5]}...") # Mostra as 5 primeiras
    except Exception as e:
        print(f"❌ Erro ao carregar usinas do banco de dados: {e}")
    return usinas_map

def processar_aba(sheet, cursor, usinas_map):
    # *** ALTERAÇÃO AQUI: Lendo cabeçalhos de data da LINHA 4 (sheet[4]) ***
    col_headers = [cell.value for cell in sheet[4]][3:] # Começa da coluna D (índice 3)
    datas = []
    for header_value in col_headers:
        print(f"[DEBUG DATE] Raw header value: {repr(header_value)}, Type: {type(header_value)}")
        if isinstance(header_value, datetime):
            datas.append(header_value.date())
            print(f"[DEBUG DATE] Converted from datetime object: {header_value.date()}")
        elif isinstance(header_value, (float, int)):
            try:
                # Usa a função from_excel do openpyxl para conversão robusta de datas
                date_obj = from_excel(header_value)
                datas.append(date_obj.date())
                print(f"[DEBUG DATE] Converted from float/int: {header_value} -> {date_obj.date()}")
            except Exception as e:
                print(f"⚠️ Erro ao converter data numérica do Excel '{header_value}': {e}")
                datas.append(None)
        elif isinstance(header_value, str):
            match = re.search(r'(\d{2}/\d{2}/\d{4})', header_value)
            if match:
                date_str = match.group(1)
                try:
                    datas.append(datetime.strptime(date_str, "%d/%m/%Y").date())
                    print(f"[DEBUG DATE] Converted from string: '{date_str}'")
                except ValueError:
                    print(f"⚠️ Erro ao converter data string '{date_str}': {e}")
                    datas.append(None)
            else:
                print(f"[DEBUG DATE] String header value '{header_value}' did not match date pattern.")
                datas.append(None)
        else:
            print(f"[DEBUG DATE] Unhandled header value type: {type(header_value)}")
            datas.append(None)

    inseridos = 0
    last_valid_usina_name_raw = None # Variável para armazenar o último nome de usina válido

    # Começa a iterar da linha 5 (min_row=5) para os dados de energia
    for linha_idx_in_sheet, linha in enumerate(sheet.iter_rows(min_row=5, values_only=True)):
        excel_row_num = linha_idx_in_sheet + 5

        if not linha or len(linha) < 3: # Precisa de pelo menos A, B, C
            print(f"[DEBUG] Linha {excel_row_num}: Linha com dados insuficientes para nome da usina. Pulando.")
            continue

        current_usina_name_cell_value = linha[2] # Coluna C é índice 2
        
        if current_usina_name_cell_value is not None and \
           (isinstance(current_usina_name_cell_value, str) and current_usina_name_cell_value.strip()):
            last_valid_usina_name_raw = current_usina_name_cell_value
        
        nome_usina_raw = last_valid_usina_name_raw

        if nome_usina_raw is None or (isinstance(nome_usina_raw, str) and not nome_usina_raw.strip()):
            print(f"[DEBUG] Linha {excel_row_num}: Nome da usina (após tratamento de mesclagem) ainda vazio ou None. Pulando.")
            continue

        print(f"[DEBUG] Linha {excel_row_num}: Nome original da usina (usado): {repr(nome_usina_raw)}")
        nome_usina_normalizado = normalizar_nome_para_comparacao(nome_usina_raw)
        print(f"[DEBUG] Linha {excel_row_num}: Nome normalizado para busca: {repr(nome_usina_normalizado)}")

        usina_id = usinas_map.get(nome_usina_normalizado)

        if usina_id is None:
            print(f"❌ Linha {excel_row_num}: Usina não encontrada no DB após normalização: {nome_usina_normalizado} (Original: {nome_usina_raw})")
            continue
        
        if len(linha) < 4: # Precisa de pelo menos A, B, C, D
            print(f"[DEBUG] Linha {excel_row_num}: Linha com dados insuficientes para energia. Pulando.")
            continue

        for i, valor in enumerate(linha[3:]): # Valores de dados começam da coluna D (índice 3)
            data = datas[i] if i < len(datas) else None
            
            if not data or valor is None or (isinstance(valor, str) and not valor.strip()):
                continue

            try:
                if isinstance(valor, datetime):
                    continue
                
                if isinstance(valor, str):
                    energia = float(valor.replace(",", "."))
                else:
                    energia = float(valor)

                print(f"📈 Linha {excel_row_num}, Data {data}: {energia} kWh na usina {nome_usina_normalizado}")
                cursor.execute(
                    '''
                    INSERT INTO "GeracaoDiaria" ("usinaId", data, "energiaKwh", ocorrencia, clima)
                    VALUES (%s, %s, %s, %s, %s)
                    ''',
                    (usina_id, data, energia, '', '')
                )
                inseridos += 1
            except ValueError as ve:
                print(f"⚠️ Linha {excel_row_num}, Data {data}: Erro de conversão de valor '{valor}' para energia: {ve}")
            except Exception as e:
                print(f"⚠️ Linha {excel_row_num}, Data {data}: Erro inesperado ao processar valor '{valor}': {e}")
    return inseridos

def contar_registros(cursor):
    cursor.execute('SELECT COUNT(*) FROM "GeracaoDiaria"')
    return cursor.fetchone()[0]

def importar_planilhas():
    conn = conectar()
    cursor = conn.cursor()
    
    usinas_map = carregar_usinas_do_db(cursor)
    if not usinas_map:
        print("❌ Não foi possível carregar as usinas do banco de dados. Abortando importação.")
        cursor.close()
        conn.close()
        return

    try:
        wb = openpyxl.load_workbook(CAMINHO_ARQUIVO)
        for nome_aba in wb.sheetnames:
            meses_regex = "(Janeiro|Fevereiro|Março|Abril|Maio|Junho|Julho|Agosto|Setembro|Outubro|Novembro|Dezembro)"
            padrao_mes_ano = rf"^{meses_regex}\s?\d{{2,4}}$"
            padrao_geracao_mm_yyyy = r"^Geração - \d{2}_\d{4}$"

            if not (re.match(padrao_mes_ano, nome_aba, re.IGNORECASE) or \
                    re.match(padrao_geracao_mm_yyyy, nome_aba)):
                print(f"\nℹ️ Pulando aba '{nome_aba}' (não corresponde aos padrões esperados).")
                continue

            sheet = wb[nome_aba]
            print(f"\n📄 Processando aba: {nome_aba} ({sheet.max_row - 1} linhas de dados esperadas)")
            try:
                antes = contar_registros(cursor)
                inseridos = processar_aba(sheet, cursor, usinas_map)
                depois = contar_registros(cursor)
                print(f"✅ Inseridos: {inseridos} | Total antes: {antes} → depois: {depois}")
            except Exception as e:
                print(f"❌ Erro na aba {nome_aba}: {e}")
    except FileNotFoundError:
        print(f"Erro: O arquivo '{CAMINHO_ARQUIVO}' não foi encontrado. Certifique-se de que a planilha.xlsx está no mesmo diretório do script.")
    except Exception as e:
        print(f"Erro geral ao importar planilhas: {e}")
    finally:
        cursor.close()
        conn.close()

if __name__ == "__main__":
    importar_planilhas()
